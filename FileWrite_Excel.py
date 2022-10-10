import openpyxl

excel_path = "E:\\kk\\sample1.xlsx" # Path of the workbook

open_workbook = openpyxl.load_workbook(excel_path) # Open workbook
open_worksheet =open_workbook["Sheet1"] # Open worksheet

# Note: Row and Column count will start from 1
# Maximum Row
maxrow = open_worksheet.max_row
print(maxrow)

# Maximum Column
maxcolumn = open_worksheet.max_column
# print(maxcolumn)
