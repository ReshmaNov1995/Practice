# Fetching Data from the DB & write it in the Excel

import pypyodbc
import openpyxl

# Connecting DB
dbcredentials = "Driver={SQL Server};Server=AGL113\SQLEXPRESS2017;Database=hylands_19May;UID=sa;PWD=admin@123"
dbconnect = pypyodbc.connect(dbcredentials)
cursor = dbconnect.cursor()

# Executing the Query & Fetching the Data
# query = "select semail from users"
query = "select * from users"
execute = cursor.execute(query)
result = cursor.fetchall()
# print(result)

# Opening Excel Sheet
excel_path = "E:\\kk\\sample2.xlsx"
open_workbook = openpyxl.load_workbook(excel_path)
open_worksheet = open_workbook["comparison"]

a = open_worksheet.max_column
# print(a)

# Writing Data from DB to Excel as Columnwise
row = 1
column = 1

for i in result:
    n = str(i)
    print(n)
    open_worksheet.cell(row, column).value = n
    column += 1
    open_workbook.save(excel_path)
    open_workbook.close()

# Writing Data from DB to Excel as Rowwise
column = 1
row = 1

for j in result:
    m = str(j)
    print(m)
    open_worksheet.cell(row, column).value = m
    row += 1
    open_workbook.save(excel_path)
    open_workbook.close()


# Writing Data from DB as a table in Excel

row = 1

for i in result:
    column = 1
    for j in i:
        open_worksheet.cell(row, column).value = j
        open_workbook.save(excel_path)
        open_workbook.close()
        column += 1

    row += 1