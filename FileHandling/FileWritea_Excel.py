import openpyxl

excel_path="E:\\KRN Files\\Python\\R Youtube link\\Part-14\\write.xlsx"
excel_workbook=openpyxl.load_workbook(excel_path)
#method-1
excel_worksheet=excel_workbook["Sheet1"]
#method-2
# excel_worksheet=excel_workbook.active#antha workbook la entha sheet active la iruko athu eduthukum if antha workbook la multiple sheet irunda method-1 tan follow pananum
excel_max_row=excel_worksheet.max_row
excel_max_column=excel_worksheet.max_column

excel_worksheet.cell(1,1).value="ate01"
excel_worksheet.cell(2,1).value="ate01"
excel_worksheet.cell(3,1).value="ate01"
excel_worksheet.cell(4,1).value="ate01"
excel_worksheet.cell(5,1).value="ate01"
excel_worksheet.cell(6,1).value="ate01"


excel_worksheet.cell(1,2).value="1"
excel_worksheet.cell(2,2).value="1"
excel_worksheet.cell(3,2).value="1"
excel_worksheet.cell(4,2).value="1"
excel_worksheet.cell(5,2).value="1"
excel_worksheet.cell(6,2).value="1"

excel_workbook.save(excel_path)