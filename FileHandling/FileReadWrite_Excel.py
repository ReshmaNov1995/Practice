import openpyxl

excel_path="E:\\kk\\sample1.xlsx"
excel_workbook=openpyxl.load_workbook(excel_path)

#method-1
# excel_worksheet=excel_workbook["Sheet1"]

#method-2
excel_worksheet=excel_workbook.active #antha workbook la entha sheet active la iruko athu eduthukum if antha workbook la multiple sheet irunda method-1 tan follow pananum
excel_max_row=excel_worksheet.max_row
excel_max_column=excel_worksheet.max_column


for i in range(1,excel_max_row+1):
    for j in range(1,excel_max_column+1):
        excel_values=excel_worksheet.cell(i,j).value
        print(excel_values,end="    ")
    print()


