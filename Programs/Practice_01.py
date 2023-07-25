"""
Step-1:
Header in a list
Step-2:
Only data in a list
Step-3:
dict = {'id':'TC001', 'TC002', 'TC003'} and so on
"""

import openpyxl

Excel_Path = "E:\\dummy\\test_datas.xlsx"
workbook = openpyxl.load_workbook(Excel_Path)
worksheet = workbook["login"]

row = worksheet.max_row
column = worksheet.max_column

header_list = []
data_list = []


for j in range(1, column+1):
    data = worksheet.cell(1,j).value
    header_list.append(data)
print(header_list)

for i in range(2, row+1):
    list1 = []
    for j in range(1, column+1):
        data = worksheet.cell(i,j).value
        list1.append(data)
    data_list.append(list1)
print(data_list)

dict1 = {}
for i in range(len(header_list)):
    for j in range(len(data_list)):
        if header_list[j] not in dict1:
            dict1[header_list[j]] = [(data_list[i])[j]]
        else:
            dict1[header_list[j]].append((data_list[i])[j])
print(dict1)



for i in range(len(dict1)):
    if (dict1["ID"])[i] == 'TC001':
        print((dict1['User Name'])[i])
        print((dict1['Password'])[i])






